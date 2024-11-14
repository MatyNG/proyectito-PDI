import pandas as pd
import tkinter as tk
from tkinter import messagebox
import os
import configparser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Cargar configuración desde el archivo INI
config = configparser.ConfigParser()
config.read('config.ini')

# Leer parámetros configurables
titulo_ventana = config['config']['titulo_ventana']
archivo_excel = os.path.join(os.getcwd(), config['config']['archivo_excel'])

# Función para guardar los datos en un archivo Excel
def guardar_datos():
    nombre = entry_nombre.get()
    apellido = entry_apellido.get()
    edad = entry_edad.get()
    telefono = entry_telefono.get()
    email = entry_email.get()
    nacionalidad = entry_nacionalidad.get()

    # Crear un DataFrame con los datos
    df = pd.DataFrame({
        'Nombre': [nombre],
        'Apellido': [apellido],
        'Edad': [edad],
        'Teléfono': [telefono],
        'Email': [email],
        'Nacionalidad': [nacionalidad]
    })

    try:
        if os.path.exists(archivo_excel):
            # Si el archivo existe, carga el libro de trabajo
            book = load_workbook(archivo_excel)

            # Verifica si la hoja 'Datos' existe, si no, la crea
            if 'Datos' in book.sheetnames:
                sheet = book['Datos']
            else:
                sheet = book.create_sheet('Datos')
                # Agregar los encabezados
                for col_num, column_title in enumerate(df.columns, 1):
                    sheet.cell(row=1, column=col_num, value=column_title)

            # Agregar los datos del DataFrame al final de la hoja
            for r in dataframe_to_rows(df, index=False, header=False):
                sheet.append(r)

            # Guardar los cambios en el archivo Excel
            book.save(archivo_excel)
            messagebox.showinfo("Éxito", "Datos guardados correctamente.")

        else:
            # Si el archivo no existe, crea uno nuevo
            with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos', index=False)

            messagebox.showinfo("Éxito", "Archivo creado y datos guardados correctamente.")

        # Limpiar los campos de entrada
        entry_nombre.delete(0, tk.END)
        entry_apellido.delete(0, tk.END)
        entry_edad.delete(0, tk.END)
        entry_telefono.delete(0, tk.END)
        entry_email.delete(0, tk.END)
        entry_nacionalidad.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar datos: {e}")

# Crear la ventana principal
root = tk.Tk()
root.title(titulo_ventana)

# Crear etiquetas y entradas
tk.Label(root, text="Nombre:").grid(row=0, column=0, padx=10, pady=10)
entry_nombre = tk.Entry(root)
entry_nombre.grid(row=0, column=1, padx=10, pady=10)

tk.Label(root, text="Apellido:").grid(row=1, column=0, padx=10, pady=10)
entry_apellido = tk.Entry(root)
entry_apellido.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Edad:").grid(row=2, column=0, padx=10, pady=10)
entry_edad = tk.Entry(root)
entry_edad.grid(row=2, column=1, padx=10, pady=10)

tk.Label(root, text="Teléfono:").grid(row=3, column=0, padx=10, pady=10)
entry_telefono = tk.Entry(root)
entry_telefono.grid(row=3, column=1, padx=10, pady=10)

tk.Label(root, text="Email:").grid(row=4, column=0, padx=10, pady=10)
entry_email = tk.Entry(root)
entry_email.grid(row=4, column=1, padx=10, pady=10)

tk.Label(root, text="Nacionalidad:").grid(row=5, column=0, padx=10, pady=10)
entry_nacionalidad = tk.Entry(root)
entry_nacionalidad.grid(row=5, column=1, padx=10, pady=10)

# Botón para guardar datos
tk.Button(root, text="Guardar", command=guardar_datos).grid(row=6, column=0, columnspan=2, pady=20)

# Ejecutar el bucle principal de la interfaz
root.mainloop()
